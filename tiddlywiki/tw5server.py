#!/usr/bin/python3
# Copyright (C) 2018-2021, 2025  macmarrum (at) outlook (dot) ie
# SPDX-License-Identifier: GPL-3.0-or-later
import contextlib
import http.server  # SimpleHTTPServer in py2
import os
import re
import socketserver  # SocketServer in py2
import sys
import time
import zipfile
from http import HTTPStatus
from pathlib import Path
from typing import Set

me = Path(__file__).absolute()
UTF8 = 'UTF-8'

try:
    from xls2tw import XlsToTableConverter, Flavor
    import openpyxl

    to_tw5 = XlsToTableConverter(Flavor.TW5)
    to_jira = XlsToTableConverter(Flavor.JIRA)
    xlsx_to_tw5_enabled = True
except ModuleNotFoundError:
    xlsx_to_tw5_enabled = False


def date_time_string():
    return time.strftime("%Y-%m-%d, %a %H:%M:%S")


def load_allowed_client_addresses(allowed_client_addresses_path: Path) -> Set[str]:
    allowed_client_addresses = {'127.0.0.1'}
    with contextlib.suppress(FileNotFoundError):
        for line in allowed_client_addresses_path.open(encoding=UTF8):
            line = line.strip()
            if line and not line.startswith('#'):
                allowed_client_addresses.add(line)
    return allowed_client_addresses


class MySimpleHTTPRequestHandler(http.server.SimpleHTTPRequestHandler):
    ALLOWED_CLIENT_ADDRESSES_PATH = me.parent / 'allowed-client-addresses.txt'
    ALLOWED_CLIENT_ADDRESSES = load_allowed_client_addresses(ALLOWED_CLIENT_ADDRESSES_PATH)
    TO_TW5_TABLE = '?to-tw5-table'
    TO_JIRA_TABLE = '?to-jira-table'
    STARTFILE = '?startfile'
    RX_TIMESTAMP = re.compile(r'[?&]timestamp=\d{13}$')
    MAX_COL = 12
    MAX_ROW = 70

    @property
    def is_for_tw5(self):
        return self.TO_TW5_TABLE in self._original_path

    @property
    def is_for_jira(self):
        return self.TO_JIRA_TABLE in self._original_path

    def handle_one_request(self):
        """macmarrum: copied from SimpleHTTPRequestHandler.handle_one_request()
        and extended with allowed-client-addresses

        Handle a single HTTP request.

        You normally don't need to override this method; see the class
        __doc__ string for information on how to handle specific HTTP
        commands such as GET and POST.

        """
        try:
            self.raw_requestline = self.rfile.readline(65537)
            if len(self.raw_requestline) > 65536:
                self.requestline = ''
                self.request_version = ''
                self.command = ''
                self.send_error(HTTPStatus.REQUEST_URI_TOO_LONG)
                return
            if not self.raw_requestline:
                self.close_connection = True
                return
            if not self.parse_request():
                # An error code has been sent, just exit
                return
            mname = 'do_' + self.command
            if not hasattr(self, mname):
                self.send_error(
                    HTTPStatus.NOT_IMPLEMENTED,
                    "Unsupported method (%r)" % self.command)
                return
            # <macmarrum>
            ip = self.client_address[0]
            if ip not in self.ALLOWED_CLIENT_ADDRESSES:
                print(f"** {ip} not in {self.ALLOWED_CLIENT_ADDRESSES_PATH.name}: {self.ALLOWED_CLIENT_ADDRESSES}")
                self.send_error(HTTPStatus.FORBIDDEN, 'Forbidden')
            else:
                # </macmarrum>
                method = getattr(self, mname)
                method()
            self.wfile.flush()  # actually send the response if not already done.
        except TimeoutError as e:
            # a read or a write timed out.  Discard this connection
            self.log_error("Request timed out: %r", e)
            self.close_connection = True
            return

    def log_message(self, format, *args):
        msg = f"{self.client_address[0]} - - [{date_time_string()}] {format % args}\n"
        sys.stderr.write(msg)
        with open(LOG_FILE, 'a') as ouf:
            ouf.write(msg)

    def _convert_to_table(self, path, convert_func):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        if ws.max_column > self.MAX_COL or ws.max_row > self.MAX_ROW:
            return f"""|{os.path.basename(path)}<br/> has too many columns and/or rows|<|<|h
|Columns | {ws.max_column}|{self.MAX_COL:>2} is the limit |
|Rows | {ws.max_row}|{self.MAX_ROW:>2} is the limit |
""".encode(UTF8)
        else:
            table_markdown = convert_func(ws)
            if self.is_for_tw5:
                link_to_tw5_table = f"[ext[text:tw5|{self._original_path}]]"
                path__to_jira_table = self._original_path.replace(self.TO_TW5_TABLE, self.TO_JIRA_TABLE)
                link_to_jira_table = f"[ext[text:jira|{path__to_jira_table}]]"
                body = f"|[ext[{self.path}{self.STARTFILE}]] {link_to_tw5_table} {link_to_jira_table}|c\n{table_markdown}"
            else:
                # jira does not need links in the footer
                #  xhr-text will send only TO_TW5_TABLE in its request
                body = table_markdown
            return body.encode(UTF8)

    def _remove_timestamp_from_path(self):
        self._original_path = self.path
        self.path = self.RX_TIMESTAMP.sub('', self.path)

    def _remove_to_table_from_path(self):
        self.path = self.path.replace(self.TO_TW5_TABLE, '')

    def _convert_to_table_and_send_head(self):
        self._remove_to_table_from_path()
        path = super().translate_path(self.path)
        try:
            path_st = os.stat(path)
            if self.is_for_tw5:
                convert_func = to_tw5.convert
            elif self.is_for_jira:
                convert_func = to_jira.convert
            else:
                raise ValueError(f"Unknown table type: {self._original_path}")
            bytes_msg = self._convert_to_table(path, convert_func)
            super().send_response(HTTPStatus.OK)
            super().send_header('Content-type', 'text/plain; charset="utf-8"')
            super().send_header('Content-Length', str(len(bytes_msg)))
            super().send_header('Last-Modified', super().date_time_string(int(path_st.st_mtime)))
            super().end_headers()
            return bytes_msg
        except (FileNotFoundError, OSError):
            self.send_error(HTTPStatus.NOT_FOUND, f"File not found: `{path}'")  # 404
            return None

    def do_GET(self):
        self.log_message(f'"{self.requestline}" (req)')
        self._remove_timestamp_from_path()
        if xlsx_to_tw5_enabled and (self.is_for_tw5 or self.is_for_jira):
            text = self._convert_to_table_and_send_head()
            if text:
                try:
                    self.wfile.write(text)
                except:
                    print(f":: {text}")
                    raise
        elif self.STARTFILE in self.path:
            super().send_response(HTTPStatus.NO_CONTENT)  # 204
            super().end_headers()
            self._remove_to_table_from_path()
            translated_path = super().translate_path(self.path)
            if sys.platform == 'win32':
                os.startfile(translated_path)
            elif sys.platform.startswith('linux'):
                import subprocess
                subprocess.call(['xdg-open', translated_path])
        else:
            super().do_GET()

    def do_HEAD(self):
        self._remove_timestamp_from_path()
        if xlsx_to_tw5_enabled and (self.is_for_tw5 or self.is_for_jira):
            self._convert_to_table_and_send_head()
        else:
            super().do_HEAD()

    def do_PUT(self):
        self.log_message(f'"{self.requestline}" "Content-Length: {self.headers.get("Content-Length")}" (req)')
        path = super().translate_path(self.path)
        resolved_path = Path(path).resolve()
        resolved_doc_root = DOC_ROOT.resolve()
        if not (resolved_path == resolved_doc_root or resolved_doc_root in resolved_path.parents):
            super().send_error(HTTPStatus.FORBIDDEN, "Forbidden: Attempted write outside document root.")
            return
        size = int(self.headers.get('Content-Length', None))
        data = self.rfile.read(size)
        with open(path, 'wb') as ouf:
            ouf.write(data)
        self._add_data_to_zipfile(path, data)
        super().send_response(HTTPStatus.OK)
        super().end_headers()

    @staticmethod
    def _add_data_to_zipfile(path, data: bytes):
        path = Path(path)
        zip_path = BACKUP_DIR / f"{path.stem}-tw5.zip"
        with zipfile.ZipFile(zip_path, 'a', zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
            zf.writestr(f"{time.strftime('%Y-%m-%d_%H·%M·%S')}~{path.name}", data)

    def do_OPTIONS(self):
        self.log_message(f'"{self.requestline}" (req)')
        super().send_response(HTTPStatus.OK)  # TiddlyWiki.html fails to call PUT when 204 is returned
        super().send_header('Allow', 'GET,HEAD,OPTIONS,PUT')
        super().send_header('x-api-access-type', 'file')
        super().send_header('dav', 'tw5/put')
        super().end_headers()


class MyTCPServer(socketserver.TCPServer):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    # @override
    def handle_error(self, request, client_address):
        """Overrides socketserver.BaseServer.handle_error(self, request, client_address)"""
        print(client_address, sys.exc_info()[1], file=sys.stderr)


PORT = 8000
HOST = '127.0.0.1' if MySimpleHTTPRequestHandler.ALLOWED_CLIENT_ADDRESSES == {'127.0.0.1'} else '0.0.0.0'
WORK_DIR = me.parent
DOC_ROOT = WORK_DIR
LOG_FILE = WORK_DIR / f"tw5-server-{PORT}.log"
BACKUP_DIR = WORK_DIR

os.chdir(DOC_ROOT)
httpd = MyTCPServer((HOST, PORT), MySimpleHTTPRequestHandler)
ip, port = httpd.server_address
date_time = date_time_string()
print(f":: Python {sys.version}")
print(f":: {date_time} -- serving '{os.getcwd()}' at {ip} {port}")
print(f":: {date_time} -- WORK_DIR: {WORK_DIR}")
print(f":: {date_time} -- allowed client addresses: {MySimpleHTTPRequestHandler.ALLOWED_CLIENT_ADDRESSES}")
httpd.serve_forever()
